import openpyxl
import sys
from pathlib import Path
from openpyxl.styles import Alignment, Font

class XlsxDataReplace:
    dataSets = []

    def __init__(self, settings, db):
        self.settings = settings
        self.db = db
    
    def run_script(self, params):
        filePath = Path(params[0])
        filePath = filePath.parent / (filePath.stem + '.xlsx')

        if not filePath.is_file():
            print("File not found: " + str(filePath))
            sys.exit(1)
        
        workbook = openpyxl.load_workbook(filePath)
        
        for dataSet in self.dataSets:
            worksheet = workbook.get_sheet_by_name(dataSet['sheetName'])

            results = self.db.query(dataSet['sql'])

            lastRow = dataSet.get('lastRow', worksheet.max_row)
            
            lastColumn = dataSet['firstColumn'] + len(results[0])           

            for row in range(dataSet['firstRow'], lastRow+1):
                for col in range(dataSet['firstColumn'], lastColumn):
                    cell = worksheet.cell(column=col, row=row)
                    cell.value = None
            
            #write rows
            for rowIndex, resultRow in enumerate(results):
                for colIndex, value in enumerate(resultRow):
                    cell = worksheet.cell(column=dataSet['firstColumn'] + colIndex, row=dataSet['firstRow'] + rowIndex)
                    cell.value = resultRow[colIndex]
            
        try:
            workbook.save(filePath)
        except PermissionError:
            print("Uable to save. Permission denied.")

