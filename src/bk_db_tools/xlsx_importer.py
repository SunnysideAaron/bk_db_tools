import openpyxl
from .importer import Importer

class XlsxImporter(Importer): 
    extension = "*.xlsx"
        
    def import_file(self, filePath):
        workbook = openpyxl.load_workbook(filePath, data_only=True)

        for sheetName in workbook.sheetnames:
            # don't import sheets marked to ignore.
            if sheetName.endswith("_ignore"):
                continue
                
            worksheet = workbook.get_sheet_by_name(sheetName)
            lastRow = worksheet.max_row + 1
            lastColumn = worksheet.max_column + 1
    
            columnNames = []
            ignoreColumns = []
            for i in range(1, lastColumn):
                columnName = worksheet.cell(row = 1, column = i).value
                
                if columnName.endswith("_ignore"):
                    ignoreColumns.append(i)
                    continue
                
                columnNames.append(columnName)
                
            sql = self.insert_table_str(sheetName, columnNames)
            
            msgcount = 0
            for r in range(2, lastRow):
                params = []
                for c in range(1, lastColumn):
                    if c in ignoreColumns:
                        continue
                    params.append(worksheet.cell(row = r, column = c).value)
                
                if r == 2:
                    self.create_table(sheetName, columnNames, params)
                    
                self.db.execute(sql, params)    
                
                self.db.commit()
                
                msgcount += 1
                if msgcount > 500:
                    print ('inserted ' + str(r) + ' rows into ' + sheetName)
                    msgcount = 0