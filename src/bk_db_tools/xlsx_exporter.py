import openpyxl
import sys
from pathlib import Path
from openpyxl.styles import Alignment, Font

class XlsxExporter: 
    def __init__(self, settings, db):
        self.settings = settings
        self.db = db
        
    def export_tables(self, params):
        sql = """
        SELECT name
        FROM sqlite_master 
        WHERE (type = 'table'
          OR type = 'view')
        AND name like ?
        ORDER BY name;
        """
        
        rows = self.db.query(sql, [params])

        if not rows:
            print("No tables found")
            return

        for row in rows:
            tableName = row['name']
            
            sql = f'SELECT * FROM {tableName};'
            rows = self.db.query(sql)
            
            self.export_rows(self.settings.xlsxDataDump, rows, tableName)
            
            print("Exported " + tableName + " to xlsx data dump file.")

    def delete_all_sheets(self, workbookPath):
        if Path(workbookPath).is_file():
            workbook = openpyxl.load_workbook(workbookPath)
        else:
            workbook = openpyxl.Workbook()    

        for sheetName in workbook.sheetnames:
            workbook.remove(workbook[sheetName])

        #we must have at least 1 sheet to save
        workbook.create_sheet("Sheet")
        
        #save workbook
        try:
            workbook.save(workbookPath)
        except PermissionError:
            print("Uable to save. Permission denied.")    

    def export_rows(self, workbookPath, rows, worksheetName = "Sheet"):
        if Path(workbookPath).is_file():
            workbook = openpyxl.load_workbook(workbookPath)
        else:
            workbook = openpyxl.Workbook()
        
        #remove sheet (delete) if it already exists.
        if worksheetName in workbook.sheetnames:
            workbook.remove(workbook[worksheetName])

        worksheet = workbook.create_sheet(worksheetName)

        if rows:
            #write header
            column = 1
            for key in dict(rows[0]).keys():
                worksheet.cell(column=column, row=1, value=key)
                column = column + 1

            #write rows
            for rowIndex, row in enumerate(rows):
                for colIndex, value in enumerate(row):
                    cell = worksheet.cell(column=colIndex + 1, row=rowIndex + 2)
                    cell.value = row[colIndex]
                    cell.alignment = Alignment(wrapText=True)

        #basic formating
        bold_font = Font(bold=True)

        for cell in worksheet["1:1"]:
            cell.font = bold_font
    
        worksheet.freeze_panes = "A2"
        
        #adjust column widths
        dims = {}
        totalColumnMax = 70 #So long text doesn't make super long columns
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    currentColumnMax = dims.get(cell.column_letter, 0)
                    cellLength = len(str(cell.value))
                    dims[cell.column_letter] = min((totalColumnMax, max((currentColumnMax, cellLength))))    
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value * 1.1 #slightly larger for whatever your font might be.

        #save workbook
        try:
            workbook.save(workbookPath)
        except PermissionError:
            print("Uable to save. Permission denied.")
